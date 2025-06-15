import os
import tempfile
import logging
import jwt
from datetime import datetime
from functools import wraps
from flask import Flask, request, send_file, jsonify, abort, make_response
import pandas as pd
import json
from werkzeug.utils import secure_filename
from werkzeug.exceptions import RequestEntityTooLarge

app = Flask(__name__)

# JWT Configuration
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'dev-secret-key-change-in-production')
app.config['JWT_ALGORITHM'] = 'HS256'

# Configurable limits
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50 MB max upload
ALLOWED_EXTENSIONS = {'.csv', '.xls', '.xlsx', '.ods', '.json'}
ALLOWED_MIME_TYPES = {
    '.csv': ['text/csv', 'application/vnd.ms-excel'],
    '.xls': ['application/vnd.ms-excel'],
    '.xlsx': ['application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'],
    '.ods': ['application/vnd.oasis.opendocument.spreadsheet'],
    '.json': ['application/json', 'text/json']
}

# Logging setup
logging.basicConfig(level=logging.INFO, format='%(asctime)s %(levelname)s %(message)s')
logger = logging.getLogger(__name__)

def allowed_file(filename):
    ext = os.path.splitext(filename)[1].lower()
    return ext in ALLOWED_EXTENSIONS

def allowed_mime(mimetype, ext):
    return mimetype in ALLOWED_MIME_TYPES.get(ext, [])

import hashlib

# JWT Authorization middleware
def token_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        token = None
        auth_header = request.headers.get('Authorization')
        
        # Check if Authorization header exists and has correct format
        if auth_header and auth_header.startswith('Bearer '):
            token = auth_header.split(' ')[1]
        
        # If running in development mode, allow requests without token
        if os.environ.get('FLASK_ENV') == 'development' and not token:
            logger.warning('No token provided, but allowing access in development mode')
            return f(*args, **kwargs)
            
        if not token:
            logger.warning('No token provided in request')
            return jsonify({'message': 'Authentication token is missing'}), 401
        
        try:
            # Verify and decode the token
            jwt.decode(token, app.config['SECRET_KEY'], algorithms=[app.config['JWT_ALGORITHM']])
        except jwt.ExpiredSignatureError:
            return jsonify({'message': 'Token has expired'}), 401
        except jwt.InvalidTokenError:
            return jsonify({'message': 'Invalid token'}), 401
            
        return f(*args, **kwargs)
    
    return decorated

def strip_formulas(df):
    # Remove formulas from all string cells (cells starting with '=')
    def clean_cell(val):
        if isinstance(val, str) and val.strip().startswith('='):
            # Just remove the formula prefix but keep the rest of the value
            return val.strip()[1:]
        return val
    
    # Use apply instead of deprecated applymap
    logger.info(f'DataFrame before formulas stripped - shape: {df.shape}, columns: {list(df.columns)}')
    logger.info(f'Sample of first few rows before cleaning: {df.head(2).to_dict()}')
    
    # Apply the clean_cell function to each column that contains string data
    for col in df.columns:
        if df[col].dtype == 'object':
            df[col] = df[col].map(clean_cell)
    
    logger.info(f'DataFrame after formulas stripped - shape: {df.shape}, columns: {list(df.columns)}')
    logger.info(f'Sample of first few rows after cleaning: {df.head(2).to_dict()}')
    
    return df

def convert_to_xlsx(input_path, output_path, ext):
    logger.info(f'Starting conversion from {ext} to XLSX')
    logger.info(f'Input file: {input_path}, Output file: {output_path}')
    
    try:
        file_size = os.path.getsize(input_path)
        logger.info(f'Input file size: {file_size} bytes')
    except Exception as e:
        logger.error(f'Error checking input file size: {str(e)}')
    
    # XLS files need special handling for proper conversion
    if ext == '.xls':
        logger.info('Special handling for XLS file - using direct cell-by-cell copy approach')
        try:
            # Use xlrd for reading XLS
            import xlrd
            from openpyxl import Workbook
            
            # Open the XLS workbook
            logger.info(f'Opening XLS file: {input_path}')
            try:
                wb = xlrd.open_workbook(input_path)
                logger.info(f'Successfully opened XLS with xlrd')
                logger.info(f'XLS has {wb.nsheets} sheets: {wb.sheet_names()}')
            except Exception as e:
                logger.error(f'Failed to open XLS with xlrd: {str(e)}')
                raise ValueError(f'Unable to process XLS file: {str(e)}')
            
            # Create a new XLSX workbook
            logger.info('Creating new XLSX workbook with openpyxl')
            output_wb = Workbook()
            
            # Remove the default sheet created by openpyxl
            if 'Sheet' in output_wb.sheetnames:
                output_wb.remove(output_wb['Sheet'])
            
            # Copy each sheet
            for sheet_idx in range(wb.nsheets):
                sheet = wb.sheet_by_index(sheet_idx)
                logger.info(f'Processing sheet {sheet_idx+1}/{wb.nsheets}: {sheet.name} ({sheet.nrows} rows, {sheet.ncols} columns)')
                
                # Create a new sheet in the output workbook
                output_sheet = output_wb.create_sheet(title=sheet.name)
                
                # Copy all the cells
                for row_idx in range(sheet.nrows):
                    if row_idx % 1000 == 0 and row_idx > 0:
                        logger.info(f'Copied {row_idx}/{sheet.nrows} rows in sheet {sheet.name}')
                    
                    for col_idx in range(sheet.ncols):
                        # Get cell value
                        cell = sheet.cell(row_idx, col_idx)
                        value = cell.value
                        
                        # Handle formula cells - strip the formula if it starts with '='
                        if isinstance(value, str) and value.strip().startswith('='):
                            value = value.strip()[1:]  # Remove the '=' prefix
                        
                        # Write to the output sheet
                        output_sheet.cell(row=row_idx+1, column=col_idx+1).value = value
                
                logger.info(f'Successfully copied all {sheet.nrows} rows from sheet: {sheet.name}')
            
            # Save the output XLSX file
            logger.info(f'Saving XLSX file to {output_path}')
            output_wb.save(output_path)
            
            # Verify output file
            if os.path.exists(output_path):
                output_size = os.path.getsize(output_path)
                logger.info(f'Output XLSX created successfully - size: {output_size} bytes')
                return
            else:
                logger.error('Failed to create output XLSX file')
                raise IOError('Failed to create output XLSX file')
                
        except Exception as e:
            logger.exception(f'Error in direct XLS to XLSX conversion: {str(e)}')
            logger.info('Falling back to pandas conversion')
            
            # Fall back to pandas conversion if direct approach fails
            try:
                logger.info('Reading XLS with pandas as fallback method')
                df = pd.read_excel(input_path)
                logger.info(f'Successfully read XLS with pandas - shape: {df.shape}')
                
                # Strip formulas if necessary
                df = strip_formulas(df)
                
                # Write to Excel
                logger.info('Writing to XLSX with pandas')
                writer = pd.ExcelWriter(output_path, engine='openpyxl')
                df.to_excel(writer, index=False, sheet_name='Converted Data')
                writer.close()
                
                # Verify pandas output
                if os.path.exists(output_path):
                    output_size = os.path.getsize(output_path)
                    logger.info(f'Pandas fallback output created - size: {output_size} bytes')
                    return
                else:
                    logger.error('Failed to create pandas fallback output')
                    raise IOError('Failed to create pandas fallback output')
                
            except Exception as inner_e:
                logger.exception(f'Pandas fallback also failed: {str(inner_e)}')
                raise ValueError(f'All XLS conversion methods failed: {str(e)} and {str(inner_e)}')
    
    # Handle other file types with the standard pandas approach
    try:
        # Read the input file based on its extension
        if ext == '.csv':
            logger.info('Reading CSV file with pandas')
            df = pd.read_csv(input_path)
        elif ext == '.xlsx':
            logger.info('Reading XLSX file with pandas using openpyxl engine')
            df = pd.read_excel(input_path, engine='openpyxl')
        elif ext == '.ods':
            logger.info('Reading ODS file with pandas using odf engine')
            df = pd.read_excel(input_path, engine='odf')
        elif ext == '.json':
            logger.info('Reading JSON file')
            with open(input_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
                if isinstance(data, list):
                    if all(isinstance(row, dict) for row in data):
                        df = pd.DataFrame(data)
                        logger.info(f'Converted JSON array of objects to DataFrame - shape: {df.shape}')
                    elif all(isinstance(row, list) for row in data):
                        df = pd.DataFrame(data)
                        logger.info(f'Converted JSON array of arrays to DataFrame - shape: {df.shape}')
                    else:
                        logger.error('Invalid JSON format: must be array of objects or arrays')
                        raise ValueError('JSON must be array of objects or array of arrays')
                else:
                    logger.error('Invalid JSON format: must be an array')
                    raise ValueError('JSON must be an array')
        else:
            logger.error(f'Unsupported file extension: {ext}')
            raise ValueError('Unsupported file type')
        
        # Log info about the loaded DataFrame
        logger.info(f'Successfully read input file - DataFrame shape: {df.shape}')
        logger.info(f'DataFrame columns: {list(df.columns)}')
        logger.info(f'DataFrame dtypes: {df.dtypes}')
        
        # Strip formulas if needed
        df = strip_formulas(df)
        
        # Write to Excel file with additional options to ensure data integrity
        logger.info(f'Writing DataFrame to XLSX file: {output_path}')
        writer = pd.ExcelWriter(output_path, engine='openpyxl')
        df.to_excel(writer, index=False, sheet_name='Converted Data')
        writer.close()
        
        # Verify output file
        if os.path.exists(output_path):
            output_size = os.path.getsize(output_path)
            logger.info(f'Output XLSX created successfully - size: {output_size} bytes')
        else:
            logger.error('Failed to create output file!')
            raise FileNotFoundError(f'Output file not created: {output_path}')
            
    except Exception as e:
        logger.exception(f'Error in standard conversion path: {str(e)}')
        raise ValueError(f'File conversion failed: {str(e)}')


@app.errorhandler(RequestEntityTooLarge)
def handle_file_too_large(e):
    logger.warning('Upload too large')
    return jsonify({'error': 'File too large'}), 413

@app.route('/convert', methods=['POST'])
@token_required
def convert():
    logger.info('Received conversion request')
    
    # Extract user ID from token for logging, or use anonymous
    hashed_user_id = 'anonymous'
    auth_header = request.headers.get('Authorization')
    if auth_header and auth_header.startswith('Bearer '):
        token = auth_header.split(' ')[1]
        try:
            decoded = jwt.decode(token, app.config['SECRET_KEY'], algorithms=[app.config['JWT_ALGORITHM']])
            if 'sub' in decoded:
                hashed_user_id = hashlib.sha256(str(decoded['sub']).encode()).hexdigest()
        except:
            pass

    if 'file' not in request.files:
        logger.warning('No file part in request')
        return jsonify({'error': 'No file part in request. Please upload a file.'}), 400
    file = request.files['file']
    filename = secure_filename(file.filename)
    ext = os.path.splitext(filename)[1].lower()
    if not allowed_file(filename):
        logger.warning(f'Unsupported file extension: {ext}')
        return jsonify({'error': f'Unsupported file extension: {ext}. Supported: .csv, .xls, .xlsx, .ods, .json'}), 400
    if not allowed_mime(file.mimetype, ext):
        logger.warning(f'Unsupported MIME type: {file.mimetype} for extension {ext}')
        return jsonify({'error': f'Unsupported MIME type: {file.mimetype}. Try re-saving your file or contact support.'}), 400
    if request.content_length and request.content_length > 50 * 1024 * 1024:
        logger.warning('File exceeds 50MB limit')
        return jsonify({'error': 'File exceeds 50MB limit. Try splitting your file or contact support for larger uploads.'}), 413
    temp_in = tempfile.NamedTemporaryFile(delete=False, suffix=ext)
    temp_out = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    try:
        file.save(temp_in.name)
        convert_to_xlsx(temp_in.name, temp_out.name, ext)
        logger.info(json.dumps({
            'timestamp': datetime.utcnow().isoformat(),
            'event': 'conversion',
            'user': hashed_user_id,
            'filename_hash': hashlib.sha256(filename.encode()).hexdigest(),
            'filetype': ext,
            'filesize': request.content_length,
            'status': 'success',
            'blockchain_anchor': None  # Placeholder for future Hedera integration
        }))
    except Exception as e:
        logger.exception('Conversion failed')
        logger.info(json.dumps({
            'timestamp': datetime.utcnow().isoformat(),
            'event': 'conversion',
            'user': hashed_user_id,
            'filename_hash': hashlib.sha256(filename.encode()).hexdigest(),
            'filetype': ext,
            'filesize': request.content_length,
            'status': 'failure',
            'error': str(e),
            'blockchain_anchor': None
        }))
        return jsonify({'error': f'Conversion failed: {str(e)}. Try saving as CSV or contact support.'}), 400
    finally:
        temp_in.close()
        if os.path.exists(temp_in.name):
            os.unlink(temp_in.name)
    logger.info(f'Successfully converted {filename} to xlsx')
    response = send_file(temp_out.name, as_attachment=True, download_name='converted.xlsx')
    def cleanup():
        if os.path.exists(temp_out.name):
            os.unlink(temp_out.name)
    response.call_on_close(cleanup)
    return response

@app.route('/health', methods=['GET'])
def health():
    return jsonify({'status': 'ok'})

@app.route('/', methods=['GET'])
def docs():
    return jsonify({
        'service': 'XLSX Conversion Microservice',
        'endpoints': {
            '/convert': {
                'method': 'POST',
                'description': 'Convert csv, xls, xlsx, ods, or json to xlsx',
                'request': 'multipart/form-data, field name: file',
                'response': 'xlsx file as attachment',
                'errors': ['File too large', 'Unsupported file type', 'Conversion error']
            },
            '/health': {
                'method': 'GET',
                'description': 'Health check',
                'response': '{"status": "ok"}'
            }
        }
    })

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5001)



@app.route('/health', methods=['GET'])
def health():
    return jsonify({'status': 'ok'})

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5001)
